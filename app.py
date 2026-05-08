from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = "movies.xlsx"

# Create Excel file if not exists
if not os.path.exists(EXCEL_FILE):

    wb = Workbook()
    ws = wb.active
    ws.title = "Movies"

    ws.append([
        "Movie Name",
        "Director",
        "Year",
        "Genre",
        "Poster",
        "Favorite"
    ])

    wb.save(EXCEL_FILE)


@app.route("/")
def index():

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    movies = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        movies.append({
            "movie": row[0],
            "director": row[1],
            "year": row[2],
            "genre": row[3],
            "poster": row[4],
            "favorite": row[5]
        })

    # Sort movies by year
    movies = sorted(movies, key=lambda x: x["year"], reverse=True)

    return render_template("index.html", movies=movies)


@app.route("/add", methods=["POST"])
def add_movie():

    movie = request.form["movie"]
    director = request.form["director"]
    year = request.form["year"]
    genre = request.form["genre"]
    poster = request.form["poster"]

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([
        movie,
        director,
        year,
        genre,
        poster,
        "No"
    ])

    wb.save(EXCEL_FILE)

    return redirect("/")


@app.route("/favorite/<movie_name>")
def favorite(movie_name):

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):

        if row[0].value == movie_name:

            if row[5].value == "Yes":
                row[5].value = "No"
            else:
                row[5].value = "Yes"

    wb.save(EXCEL_FILE)

    return redirect("/")
@app.route("/delete/<movie_name>")
def delete_movie(movie_name):

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):

        if ws.cell(row=row, column=1).value == movie_name:

            ws.delete_rows(row)
            break

    wb.save(EXCEL_FILE)

    return redirect("/")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)